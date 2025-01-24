from math import radians, sin, cos, sqrt, atan2
import pandas as pd
from openpyxl import load_workbook
import openpyxl

umkreis = 10

def calculate_distance(lat1, lon1, lat2, lon2):
    # Radius der Erde in km
    radius = 6371.0

    # Umwandlung der Breiten- und Längengrade in Radianten
    lat1_rad = radians(lat1)
    lon1_rad = radians(lon1)
    lat2_rad = radians(lat2)
    lon2_rad = radians(lon2)

    # Differenzen der Breiten- und Längengrade
    dlon = lon2_rad - lon1_rad
    dlat = lat2_rad - lat1_rad

    # Haversine-Formel zur Berechnung des Abstands
    a = sin(dlat / 2)**2 + cos(lat1_rad) * cos(lat2_rad) * sin(dlon / 2)**2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))

    # Abstand in km
    distance = radius * c
    return distance

def group_coordinates(coordinates):
    grouped_coordinates = []


    for i, coord1 in enumerate(coordinates):
        # Überprüfung, ob die Koordinate bereits in einer Gruppe enthalten ist
        if any(coord1 in group for group in grouped_coordinates):
            continue

        group = [coord1]

        for j, coord2 in enumerate(coordinates):
            if i != j:
                distance = calculate_distance(coord1[0], coord1[1], coord2[0], coord2[1])
                if distance <= umkreis:
                    # Überprüfung, ob die Koordinate bereits in einer Gruppe enthalten ist
                    if any(coord2 in group for group in grouped_coordinates):
                        continue
                    group.append(coord2)

        grouped_coordinates.append(group)

    return grouped_coordinates

def select_group_center(groups):
    group_centers = []

    for group in groups:
        center = None
        max_count = 0

        for coord in group:
            count = 0

            for other_coord in group:
                distance = calculate_distance(coord[0], coord[1], other_coord[0], other_coord[1])
                if distance <= umkreis:
                    count += 1

            if count > max_count:
                max_count = count
                center = coord

        if center is not None:
            group_centers.append(center)

    return group_centers

# Beispiel-Koordinatenliste
# Lese die Koordinaten aus der Excel-Datei
# Pfad zur Excel-Datei angeben
excel_datei = 'anschlussstellen-autobahnen-und-schnellstraßen.xlsx'

# Lade die vorhandene Excel-Datei
workbook = load_workbook(excel_datei)

# Überprüfe, ob das Tabellenblatt "Server" bereits existiert
if 'Server' in workbook.sheetnames:
    #print("Blatt vorhanden")
    # Tabellenblatt auswählen
    tabellenblatt_name = "Server"
    tabellenblatt = workbook[tabellenblatt_name]

    # Tabellenblatt löschen
    workbook.remove(tabellenblatt)

    # Arbeitsmappe speichern
    workbook.save(excel_datei)

# Daten aus den Spalten "Typ", "Latitude" und "Longitude" einlesen
daten = pd.read_excel(excel_datei, usecols=['Typ', 'Latitude', 'Longitude'])

# Daten filtern basierend auf den angegebenen Wörtern in der Spalte "Typ"
#gewuenschte_woerter = ['Anschlussstelle', 'Knoten', 'Raststaette', 'Hauptfahrstreifen', 'Staatsgrenze']
#daten = daten[daten['Typ'].isin(gewuenschte_woerter)]

# Leere Zeilen löschen
daten = daten.dropna()

# Löschen der Zeilen, in denen der Wert in der Spalte "Typ" den Begriff "server" enthält
daten = daten[daten['Typ'].str.lower() != 'server']

# Koordinaten in das gewünschte Format konvertieren und ausgeben
coordinates = [(float(lat), float(lon)) for lat, lon in zip(daten['Latitude'], daten['Longitude'])]

# Gruppierung der Koordinaten
groups = group_coordinates(coordinates)

# Auswahl des Gruppenzentrums
centers = select_group_center(groups)

# Erstellen eines DataFrames für die aktualisierten Koordinaten
updated_data = pd.DataFrame(columns=['ID', 'Autobahn', 'Strassenbezeichnung', 'Anschlussstelle', 'Typ', 'Kilometer', 'Bundesland', 'Latitude', 'Longitude'])

# Generierung der ID-Werte
ids = range(1, len(centers) + 1)

# Hinzufügen der Werte in den Dataframe
for i, center in enumerate(centers):
    row = {
        'ID': ids[i],
        'Autobahn': '',
        'Strassenbezeichnung': '',
        'Anschlussstelle': '',
        'Typ': 'Server',
        'Kilometer': '',
        'Bundesland': '',
        'Latitude': str(center[0]).replace(',', '.'),
        'Longitude': str(center[1]).replace(',', '.')
    }
    updated_data.loc[len(updated_data)] = row

# Speichern des DataFrames in das neue Tabellenblatt "Server" in der Excel-Datei
with pd.ExcelWriter(excel_datei, engine='openpyxl', mode='a') as writer:
    updated_data.to_excel(writer, sheet_name='Server', index=False)

# Ausgabe der Gruppen und ihrer Zentren
for i, group in enumerate(groups):
    print(f"Gruppe {i+1}:")
    for coord in group:
        print(coord)
    print("Gruppenzentrum:", centers[i] if i < len(centers) else "N/A")
    print()
